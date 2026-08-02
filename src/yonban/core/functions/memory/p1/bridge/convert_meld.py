# convert_meld.py — meld_sch xlsx → tsv 一次性转换(Node 侧读 tsv)
# 列名自动探测但如实打印,便于人工核对(不静默猜列)
import sys, io, json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SRC = r"D:\shajiuguan\p1shiyanshi\09_P1自驱动_专项\06_资源库_词库\P1资源库\meld_sch_concreteness\Concretenss Ratings of 9877 Two Character Chinese Words.xlsx"
DST = r"D:\shajiuguan\自驱动召回\resources_derived\meld_sch_concreteness.tsv"

import openpyxl

wb = openpyxl.load_workbook(SRC, read_only=True)
ws = wb[wb.sheetnames[0]]
rows = ws.iter_rows(values_only=True)
header = [str(c) if c is not None else "" for c in next(rows)]
print("sheets:", wb.sheetnames)
print("header:", json.dumps(header, ensure_ascii=False))

# 真实表头(2026-08-01实测): Word ID / Word / Mean of Valid Ratings / SD... / No....
word_idx = next(i for i, h in enumerate(header) if h.strip().lower() == "word")
conc_idx = next(i for i, h in enumerate(header) if h.strip().lower().startswith("mean"))
print(f"word_col={header[word_idx]!r}(idx{word_idx}) conc_col={header[conc_idx]!r}(idx{conc_idx})")

# 量表方向实测(2026-08-01 确诊): 本表 1=具体 5=抽象(苹果1.259/石头1.24 vs 觉得4.042/抽象4.423),
# 与 brysbaert(5=具体, Python=4.96)相反。conc_aligned = 6 - raw 统一为 5=具体,下游只用 aligned 列。
import os
os.makedirs(os.path.dirname(DST), exist_ok=True)
n = 0
with open(DST, "w", encoding="utf-8") as f:
    f.write("word\traw_rating\tconc_aligned\n")
    for r in rows:
        if r is None or r[word_idx] is None or r[conc_idx] is None:
            continue
        try:
            raw = float(r[conc_idx])
            f.write(f"{str(r[word_idx]).strip()}\t{raw:.3f}\t{6 - raw:.3f}\n")
            n += 1
        except (ValueError, TypeError):
            continue
print("written:", n, "->", DST)

# -*- coding: utf-8 -*-
"""
datafit.py —— 数据函数验证路（v2.4, 凛倾流程图数据段:
"先输出算法函数进行匹配看看是否正确-然后再看运行数据"）

职责: AI 写 transform(rows) 函数 → 本模块受限执行——默认只对样本行试跑,
回喂[样本输入+输出]给 AI 核对; 核对通过后 full=True 全量执行, 聚合结果由 AI
写回 spec 的 chart/table（函数契约=原始数据→图表级聚合, 产物天然小, 零 solver 改动）。

安全边界（执行 AI 代码=高危, 三层）:
  1. JS 侧 owner 闸 deployGatedAllow("allowPptDataFunc", "BEILU_PPT_DATAFUNC")——server
     部署默认拒, 本地恒放行（与 allowPptPipeline 分键: 固定管线代码≠AI 任意代码, 风险不同级）
  2. 独立子进程执行（JS 起 python, 继承 timeoutMs 超时熔断/进程隔离）
  3. 受限 exec: builtins 白名单（无 open/exec/eval/compile/input）+ import 白名单
     （math/re/datetime/statistics/collections/json/itertools/functools）——禁文件/网络/子进程
数据读取: csv(零依赖 DictReader) / json(list) / xlsx(有 openpyxl 才支持, 缺→诚实报错不猜)。
rows 形态: 字典列表（表头为键, 字符串值; 类型转换是 transform 函数自己的职责——契约简单可预期）。
"""
import builtins
import csv
import json
import os

MAX_RESULT_CHARS = 8000   # 回喂结果截断（防撑爆上下文; 全量聚合结果本应远小于此）
MAX_ROWS = 200000         # 全量行数上限（防误喂巨型文件拖死子进程）

_SAFE_BUILTIN_NAMES = (
    "len", "sum", "min", "max", "sorted", "range", "enumerate", "zip", "map",
    "filter", "abs", "round", "int", "float", "str", "bool", "list", "dict",
    "set", "tuple", "any", "all", "reversed", "isinstance", "print", "repr",
    "divmod", "pow", "ValueError", "TypeError", "KeyError", "Exception",
)
_ALLOWED_MODULES = {"math", "re", "datetime", "statistics", "collections",
                    "json", "itertools", "functools"}


def _limited_import(name, *args, **kwargs):
    if name.split(".")[0] not in _ALLOWED_MODULES:
        raise ImportError("模块 %r 不在白名单 %s" % (name, sorted(_ALLOWED_MODULES)))
    return __import__(name, *args, **kwargs)


def _load_transform(code):
    g = {"__builtins__": dict(
        {k: getattr(builtins, k) for k in _SAFE_BUILTIN_NAMES},
        __import__=_limited_import)}
    exec(compile(code, "<transform>", "exec"), g)  # noqa: S102 —— 三层安全边界见模块头
    fn = g.get("transform")
    if not callable(fn):
        raise ValueError("代码里必须定义 def transform(rows): ...")
    return fn


def _read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        # utf-8-sig: Excel 导出 csv 常带 BOM
        with open(path, encoding="utf-8-sig", newline="") as f:
            return [dict(r) for r in csv.DictReader(f)][:MAX_ROWS]
    if ext == ".json":
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        if not isinstance(d, list):
            raise ValueError("json 数据文件必须是数组（每元素一行）")
        return d[:MAX_ROWS]
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("读 xlsx 需要 openpyxl（未安装）。可先另存为 csv 再来")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "col%d" % i
                   for i, h in enumerate(next(it, []) or [])]
        rows = []
        for r in it:
            rows.append({headers[i]: ("" if v is None else str(v))
                         for i, v in enumerate(r) if i < len(headers)})
            if len(rows) >= MAX_ROWS:
                break
        wb.close()
        return rows
    raise ValueError("不支持的数据文件类型 %r（支持 csv/json/xlsx）" % ext)


def run_datafit(data_path, code_path, sample_n=8, full=False):
    """返回 dict（含 error 键=失败但不 raise, 结构化回喂 AI 自纠）。"""
    out = {"file": data_path, "full": bool(full)}
    try:
        if not os.path.isfile(data_path):
            raise FileNotFoundError("数据文件不存在: %s" % data_path)
        with open(code_path, encoding="utf-8") as f:
            code = f.read()
        rows = _read_rows(data_path)
        out["row_count"] = len(rows)
        sample_n = max(1, min(int(sample_n or 8), 50))
        sample = rows[:sample_n]
        out["sample_rows"] = sample
        fn = _load_transform(code)
        result = fn(rows if full else sample)
        txt = json.dumps(result, ensure_ascii=False, default=str)
        if len(txt) > MAX_RESULT_CHARS:
            txt = txt[:MAX_RESULT_CHARS] + "…(截断)"
            out["result_truncated"] = True
        out["result"] = txt
    except Exception as e:
        out["error"] = "%s: %s" % (type(e).__name__, str(e)[:300])
    return out
